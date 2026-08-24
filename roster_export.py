#!/usr/bin/env python3
"""Выгрузка состава в табличный файл — под заявку в лигу.

Зачем. Заявку в лигу тренер набивает руками каждый сезон, хотя все эти люди у
бота уже есть: фамилия, имя, дата рождения. Переписывание сорока строк из чата
в таблицу — работа, которой быть не должно.

Формат. Отдаём xlsx, если в окружении есть openpyxl, и CSV, если нет. Выбор
делается на месте, а не настройкой: бот живёт на сервере, куда библиотеку
могут поставить или не поставить, и падать из-за этого выгрузка не должна.

CSV остаётся полноценным запасным вариантом — он открывается и Excel, и Google
Таблицами, и в заявку оттуда копируется так же. Две мелочи, без которых он
открывается плохо:

* **BOM в начале** — иначе русский Excel читает файл как cp1251 и показывает
  кракозябры;
* **точка с запятой вместо запятой** — в русской локали Excel делит строки
  именно по ней, а по запятой сваливает всё в один столбец.

Дата рождения хранится как «2001-09-22», а в заявках её пишут «22.09.2001» —
переворачиваем здесь, чтобы тренер не правил сорок ячеек руками.

ФИО тут есть, и это не противоречит юр-инварианту проекта: лист «Игроки» —
собственная таблица команды, тренер видит эти имена и в боте. Наружу файл не
уходит: он приходит в личку тому, кто нажал.
"""

from __future__ import annotations

import csv
import io
import logging
from datetime import date
from typing import Any, Dict, List, Sequence, Tuple

logger = logging.getLogger(__name__)

# Столбцы заявки: что показываем и откуда берём. Порядок — как в бланке.
COLUMNS: Tuple[Tuple[str, str], ...] = (
    ("№", "_index"),
    ("Фамилия", "surname"),
    ("Имя", "name"),
    ("Дата рождения", "birthday"),
    ("Роль", "role"),
    ("Команда", "team"),
)

SEPARATOR = ";"
BOM = "﻿"


def human_date(raw: Any) -> str:
    """«2001-09-22» → «22.09.2001». Непонятное отдаём как есть."""
    got = str(raw or "").strip()
    if len(got) == 10 and got[4] == "-" and got[7] == "-":
        year, month, day = got[:4], got[5:7], got[8:10]
        if year.isdigit() and month.isdigit() and day.isdigit():
            return f"{day}.{month}.{year}"
    return got


def rows(people: Sequence[Dict[str, Any]]) -> List[List[str]]:
    """Строки таблицы: заголовок и люди. Порядок — как пришли."""
    out = [[title for title, _ in COLUMNS]]
    for number, person in enumerate(people, start=1):
        line = []
        for _, key in COLUMNS:
            if key == "_index":
                line.append(str(number))
            elif key == "birthday":
                line.append(human_date(person.get("birthday")))
            else:
                line.append(str(person.get(key) or "").strip())
        out.append(line)
    return out


def csv_bytes(people: Sequence[Dict[str, Any]]) -> bytes:
    """Готовый файл. Пустой состав тоже отдаём — с одним заголовком."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=SEPARATOR, lineterminator="\r\n",
                        quoting=csv.QUOTE_MINIMAL)
    writer.writerows(rows(people))
    return (BOM + buf.getvalue()).encode("utf-8")


def has_xlsx() -> bool:
    """Есть ли чем собрать настоящую книгу Excel."""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return False
    return True


def _as_date(raw: Any):
    """«2001-09-22» → date. Не разобрали — None, положим текстом."""
    got = str(raw or "").strip()
    try:
        return date(int(got[:4]), int(got[5:7]), int(got[8:10]))
    except (ValueError, IndexError):
        return None


def xlsx_bytes(people: Sequence[Dict[str, Any]], title: str = "Заявка") -> bytes:
    """Книга Excel: шапка выделена, ширины подобраны, шапка не уезжает.

    Дату рождения кладём НАСТОЯЩЕЙ датой с форматом ДД.ММ.ГГГГ, а не строкой:
    так она и показывается привычно, и сортируется правильно, если тренер
    захочет переставить строки. Что не разобралось — кладём текстом, чтобы не
    потерять значение молча."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    book = Workbook()
    sheet = book.active
    # Имя листа: Excel не пускает длиннее 31 знака и часть символов.
    clean = "".join(c for c in str(title or "Заявка") if c not in "[]:*?/\\")
    sheet.title = (clean[:31] or "Заявка")

    table = rows(people)
    sheet.append(table[0])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"

    # Числа и даты кладём типами, а не текстом: иначе таблица не сортируется,
    # а Excel вешает на каждую клетку зелёный уголок «число как текст».
    born_at = [i for i, (_, key) in enumerate(COLUMNS) if key == "birthday"]
    index_at = [i for i, (_, key) in enumerate(COLUMNS) if key == "_index"]
    for person, line in zip(people, table[1:]):
        sheet.append(line)
        for i in index_at:
            cell = sheet.cell(row=sheet.max_row, column=i + 1)
            if str(cell.value).isdigit():
                cell.value = int(cell.value)
                cell.alignment = Alignment(horizontal="center")
        for i in born_at:
            got = _as_date(person.get("birthday"))
            if got:
                cell = sheet.cell(row=sheet.max_row, column=i + 1)
                cell.value = got
                cell.number_format = "DD.MM.YYYY"

    widths = [max(len(str(line[i])) for line in table) + 2
              for i in range(len(COLUMNS))]
    for i, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = min(max(width, 6), 34)

    buf = io.BytesIO()
    book.save(buf)
    return buf.getvalue()


def file_name(what: str, today: date = None, suffix: str = "csv") -> str:
    """Имя файла: «Заявка — Второй состав — 24.08.2026.xlsx».

    Дата в имени не украшение: заявку подают несколько раз за сезон, и две
    версии в загрузках без даты не различить."""
    day = (today or date.today()).strftime("%d.%m.%Y")
    clean = " ".join(str(what or "").split()) or "команда"
    # Знаки, на которых спотыкаются файловые системы и Telegram.
    for bad in '\\/:*?"<>|':
        clean = clean.replace(bad, " ")
    return f"Заявка — {' '.join(clean.split())} — {day}.{suffix}"


def build(people: Sequence[Dict[str, Any]], what: str) -> Tuple[bytes, str]:
    """Готовый файл и его имя. Книгу Excel — если есть чем, иначе CSV."""
    if has_xlsx():
        try:
            return xlsx_bytes(people, what), file_name(what, suffix="xlsx")
        except Exception:
            # Библиотека есть, но собрать не вышло — отдаём CSV, а не ошибку:
            # тренеру нужен список, а не разбирательство с форматом.
            logger.warning("Книга Excel не собралась, отдаю CSV", exc_info=True)
    return csv_bytes(people), file_name(what, suffix="csv")


def team_people() -> List[Dict[str, Any]]:
    """Все из листа «Игроки» — по алфавиту, как в самом листе."""
    import coach_payments
    return coach_payments.players()


def group_people(group_id: int) -> List[Dict[str, Any]]:
    """Состав одной группы."""
    import player_groups
    return player_groups.members(group_id)


def missing_birthday(people: Sequence[Dict[str, Any]]) -> List[str]:
    """Кто без даты рождения — заявку с пустой клеткой не примут."""
    return [str(p.get("title") or f"{p.get('surname')} {p.get('name')}").strip()
            for p in people if not str(p.get("birthday") or "").strip()]
